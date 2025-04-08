import os
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
# 获取当前脚本的文件夹路径
current_folder = os.path.dirname(os.path.abspath(__file__))

# 作者信息
def show_personal_info():
    info = '''
    作者: 卷包一乙赖灿兴
    版本: 1.0
    日期: 2025年2月26日
    电子邮件: laicx@stu.xmu.edu.cn
    联系方式: 18046307080
    代码用途: 一二区各辅料当月单耗与历史单耗对账
    描述: 该程序用于一二区各辅料当月单耗与历史单耗对账
    
    如有问题和需要变更需求请及时与我取得联系QAQ~ 
    
    点关闭后程序执行完毕...
    请于'./Auxiliary_Material_Consumption_Reconciliation/_internal'中提取输出的表格
    '''
    
    # 创建一个 Tkinter 窗口
    window = tk.Tk()
    window.title("个人信息")
    
    # 创建一个 Label 显示个人信息
    label = tk.Label(window, text=info, padx=10, pady=10)
    label.pack()
    
    # 添加一个按钮来关闭窗口
    button = tk.Button(window, text="关闭", command=window.quit)
    button.pack(pady=10)
    
    # 运行窗口
    window.mainloop()
    
# 统一文件路径
def get_file_path(filename):
    return os.path.join(current_folder, filename)

# 一、生成单耗检查表的代码
def generate_consumption_check_table():
    # 读取所有数据表
    inventory_encoding = pd.read_excel(get_file_path("对照表.xlsx"), sheet_name="存货分类", dtype={'存货分类编码': str})
    brand_table = pd.read_excel(get_file_path("对照表.xlsx"), sheet_name="烟支型号")
    filter_stick_conversion = pd.read_excel(get_file_path("对照表.xlsx"), sheet_name="烟支长度及滤棒长度")
    filter_name_conversion = pd.read_excel(get_file_path("对照表.xlsx"), sheet_name="滤棒改名")
    production_table = pd.read_excel(get_file_path("卷包月末产量表.xlsx"), skiprows=2)
    molding_production_table = pd.read_excel(get_file_path("成型.xls"), sheet_name="报表").transpose()
    area_1_auxiliary = pd.read_excel(get_file_path("物资出库汇总表.xls"), skiprows=3, dtype={'存货编码': str})

    # 2. 清洗产量表中的牌号
    production_table["成本对象"] = production_table["牌号名称"].str.replace(r"，龙岩.*?\)", ")", regex=True).replace(r"福样促试烟", "福样促试", regex=True)

    # 3. 清洗成型产量表中的牌号（参照对照表）
    rename_dict = filter_name_conversion.set_index('卷包滤棒名称')['滤棒名称'].to_dict()

    for index, row in molding_production_table.iterrows():  
        if '二区' in str(row.iloc[0]):
            molding_production_table.at[index, '区域'] = '二区'
        else:
            molding_production_table.at[index, '区域'] = '一区'

    molding_production_table['成本对象'] = molding_production_table.iloc[:, 0].map(rename_dict).fillna(molding_production_table.iloc[0])

    # 获取原来的列名列表并修改
    old_columns = molding_production_table.columns.tolist()
    old_columns[2] = '本月生产'
    molding_production_table.columns = old_columns

    # 4. 处理物资出库汇总表
    area_1_auxiliary['区域'] = area_1_auxiliary['部门'].str[-2:]
    area_1_auxiliary['部门'] = area_1_auxiliary['部门'].str[:-2]
    area_1_auxiliary['存货分类编码'] = area_1_auxiliary['存货编码'].str[:4]

    # 匹配成型产量
    molding_production_table = molding_production_table.dropna(subset=['成本对象'])
    area_1_auxiliary = pd.merge(area_1_auxiliary, molding_production_table[['成本对象', '本月生产', '区域']], on=['成本对象', '区域'], how='left')

    # 匹配卷接包产量
    production_table = production_table.dropna(subset=['成本对象'])
    columns = production_table.columns.tolist()
    index_of_target = columns.index('实际产量')
    production_area_1 = columns[index_of_target + 1]
    production_area_2 = columns[index_of_target + 2]
    area_1_auxiliary = pd.merge(area_1_auxiliary, production_table[['成本对象', f'{production_area_1}', f'{production_area_2}']], on='成本对象', how='left')

    # 匹配折算系数
    area_1_auxiliary = pd.merge(area_1_auxiliary, filter_stick_conversion[['牌号', '盘纸折算系数59', '滤棒数量折算系数120']], left_on='成本对象', right_on='牌号', how='left')

    # 匹配存货分类对照表
    area_1_auxiliary = pd.merge(area_1_auxiliary, inventory_encoding, on='存货分类编码', how='left')
    # 调整列顺序，将存货分类编码和成本要素插入到存货编码右边
    cols = list(area_1_auxiliary.columns)
    index_of_target = cols.index('存货编码')
    cols.insert(index_of_target + 1, cols.pop(cols.index('存货分类编码')))
    cols.insert(index_of_target + 2, cols.pop(cols.index('成本要素')))
    area_1_auxiliary = area_1_auxiliary[cols]

    # 匹配烟支型号对照表
    area_1_auxiliary = pd.merge(area_1_auxiliary, brand_table, left_on='成本对象', right_on='牌号', how='left')

    # 是否宽盘
    area_1_auxiliary['是否宽盘'] = area_1_auxiliary['存货名称'].astype(str).apply(lambda x: 2 if '宽盘' in x else 1)

    # 5. 清洗物资出库汇总表中的存货分类编码和成本要素
    area_1_auxiliary['存货分类编码'] = area_1_auxiliary['存货编码'].apply(
        lambda x: x[:6] if isinstance(x, str) and x[:4] == '0312' else x[:4] if isinstance(x, str) else None
    )

    # 根据“存货分类编码”与“对照表.xlsx”中的“存货分类”表进行匹配
    area_1_auxiliary['存货分类编码'] = area_1_auxiliary['存货分类编码'].apply(
        lambda x: x if x in inventory_encoding['存货分类编码'].values else ' '
    )

    # 根据“存货分类编码”映射“成本要素”
    area_1_auxiliary = pd.merge(area_1_auxiliary, inventory_encoding[['存货分类编码', '成本要素']],
                                on='存货分类编码', how='left')

    # 删除多余的“成本要素_x”和“成本要素_y”，保留“成本要素_y”并重命名
    area_1_auxiliary = area_1_auxiliary.drop(columns=['成本要素_x'])  # 删除成本要素_x
    area_1_auxiliary = area_1_auxiliary.rename(columns={'成本要素_y': '成本要素'})  # 重命名成本要素_y

    # 再次调整列顺序，确保存货分类编码和成本要素在存货编码右边
    cols = list(area_1_auxiliary.columns)
    index_of_target = cols.index('存货编码')
    cols.insert(index_of_target + 1, cols.pop(cols.index('存货分类编码')))
    cols.insert(index_of_target + 2, cols.pop(cols.index('成本要素')))
    area_1_auxiliary = area_1_auxiliary[cols]

    # 6. 定义函数：整合产量、计算折算产量_滤棒卷烟纸、计算数量_卷烟纸换算万米、计算万支单耗、计算箱单耗
    def assign_value(row):
        if row['部门'] == '成型' and row['区域'] == '一区':
            return row['本月生产']
        if row['部门'] == '成型' and row['区域'] == '二区':
            return row['本月生产']
        if row['部门'] == '卷接包' and row['区域'] == '一区':
            return row[f'{production_area_1}']
        elif row['部门'] == '卷接包' and row['区域'] == '二区':
            return row[f'{production_area_2}']
        else:
            return None

    def z_production(row):
        if row['部门'] == '成型':
            return row['产量'] * row['滤棒数量折算系数120']
        if row['部门'] == '卷接包' and row['成本要素'] == '卷烟纸':
            return row['产量'] * row['盘纸折算系数59']
        else:
            return row['产量']

    def z_fuliao(row):
        if row['部门'] == '卷接包' and row['成本要素'] == '卷烟纸':
            return row['辅数量'] * float(row['规格'][-5:-1]) * row['是否宽盘']
        else:
            return row['数量']

    def unit_wanzhi(row):
        try:
            return row['数量_卷烟纸换算米'] / row['折算产量_滤棒卷烟纸']
        except ZeroDivisionError:
            return None

    def unit_xiang(row):
        try:
            return row['数量_卷烟纸换算米'] / row['折算产量_滤棒卷烟纸'] * 5
        except ZeroDivisionError:
            return None

    # 7. 调用函数
    area_1_auxiliary['产量'] = area_1_auxiliary.apply(assign_value, axis=1)
    area_1_auxiliary['折算产量_滤棒卷烟纸'] = area_1_auxiliary.apply(z_production, axis=1)
    area_1_auxiliary['数量_卷烟纸换算米'] = area_1_auxiliary.apply(z_fuliao, axis=1)
    area_1_auxiliary['万支单耗'] = area_1_auxiliary.apply(unit_wanzhi, axis=1)
    area_1_auxiliary['箱单耗'] = area_1_auxiliary.apply(unit_xiang, axis=1)

    # 8. 简单清洗
    columns_to_drop = ['业务财务换算率', '包装数量', '单价', '金额', '辅计量单位', '本月生产', f'{production_area_1}', f'{production_area_2}', '牌号_x', '牌号_y']
    area_1_auxiliary = area_1_auxiliary.drop(columns=columns_to_drop)
    area_1_auxiliary.loc[area_1_auxiliary['成本要素'] != '卷烟纸', '盘纸折算系数59'] = None
    area_1_auxiliary.loc[area_1_auxiliary['部门'] != '成型', '滤棒数量折算系数120'] = None

    # 定义新的列顺序
    new_column_order = ['成本对象', '部门', '区域', '存货编码', '存货分类编码', '成本要素', '存货名称', '规格', '型号', '数量', '辅数量', '计量单位', '产量', '折算产量_滤棒卷烟纸', '数量_卷烟纸换算米', '万支单耗', '箱单耗', '存货分类名称', '福建生产表内名称', '烟支类型（细）', '烟支类型（粗）', '收发类别', '盘纸折算系数59', '滤棒数量折算系数120', '是否宽盘']
    area_1_auxiliary = area_1_auxiliary[new_column_order]

    # 删除“收发类别”列为“移库出库”的行
    area_1_auxiliary = area_1_auxiliary[area_1_auxiliary['收发类别'] != '移库出库']

    # 尝试将'成本对象'列转换为字符串类型并进行过滤
    area_1_auxiliary['成本对象'] = area_1_auxiliary['成本对象'].fillna('').astype(str)
    area_1_auxiliary = area_1_auxiliary[~area_1_auxiliary['成本对象'].str.contains('合计')]

    # 导出单耗检查表
    area_1_auxiliary.to_excel(get_file_path("单耗检查表.xlsx"), index=False)

# 二、调整真实消耗基础数据
def adjust_real_consumption_data():
    # 读取数据
    area_1_auxiliary = pd.read_excel(get_file_path("单耗检查表.xlsx"), dtype={'存货分类编码': str, '存货编码': str})
    real_table = pd.read_excel(get_file_path("真实消耗基础数据.xlsx"), dtype={'分类代码': str, '存货编码': str})

    # 处理数据
    area_1_auxiliary['年'] = 2025
    m = int(input('请输入本月月份数字：'))
    area_1_auxiliary['月份'] = m
    area_1_auxiliary['备注'] = None
    pd.set_option('display.max_columns', None)  # 显示完整的列
    print(area_1_auxiliary)

    # 计算真实单箱耗（折算回去）
    def unit_xiang(row):
        if row['部门'] == '成型':
            return row['万支单耗'] * row['滤棒数量折算系数120']
        if row['部门'] == '卷接包' and row['成本要素'] == '卷烟纸':
            return row['箱单耗'] * row['盘纸折算系数59']
        return row['箱单耗']

    # 成型折算单耗换算（成型纸取万支单耗）
    def z_unit_juan(row):
        if row['部门'] == '成型':
            return row['万支单耗']
        return row['箱单耗']

    # 数量整合（卷烟纸的数量要用米而不是公斤）
    def shuliang(row):
        if row['成本要素'] == '卷烟纸':
            return row['数量_卷烟纸换算米']
        return row['数量']

    # 单位整合（卷烟纸的数量要用米而不是公斤）
    def danwei(row):
        if row['成本要素'] == '卷烟纸':
            return '米'
        return row['计量单位']

    # 应用函数来计算新的列
    area_1_auxiliary['单耗'] = area_1_auxiliary.apply(unit_xiang, axis=1)
    area_1_auxiliary['折算单耗'] = area_1_auxiliary.apply(z_unit_juan, axis=1)
    area_1_auxiliary['数量2'] = area_1_auxiliary.apply(shuliang, axis=1)
    area_1_auxiliary['计量单位2'] = area_1_auxiliary.apply(danwei, axis=1)

    # 删除“成本对象”列为空的行
    area_1_auxiliary = area_1_auxiliary.dropna(subset=['成本对象'])

    # 筛选出“成本对象”列不等于“福样促试”的行，从而删除满足条件的行
    area_1_auxiliary = area_1_auxiliary[area_1_auxiliary["成本对象"] != "福样促试"]

    # 选取列，并定义新的列顺序
    new_column_order = ['年', '月份', '区域', '成本要素', '成本对象', '存货编码', '存货名称', '规格', '辅数量', '数量2', '计量单位2', '产量', '单耗', '折算产量_滤棒卷烟纸', '折算单耗', '存货分类编码', '烟支类型（细）', '烟支类型（粗）']
    area_1_auxiliary = area_1_auxiliary[new_column_order]

    # 定义新的列名
    new_column_names = ['年', '月份', '区域', '辅料类别', '成本对象', '存货编码', '存货名称', '规格', '辅数量', '数量', '计量单位', '产量（万支）', '单耗', '折算产量', '折算单耗', '分类代码', '烟支类型（细）', '烟支类型（粗）', '备注']
    
    # 使用 rename 方法重命名列
    area_1_auxiliary = area_1_auxiliary.rename(columns=dict(zip(area_1_auxiliary.columns, new_column_names)))

    # 剔除"辅料类别"为空值的行
    area_1_auxiliary = area_1_auxiliary.dropna(subset=['辅料类别'])

    # 追加数据框 df1 和 df2 到新表
    real_table = pd.concat([real_table, area_1_auxiliary], ignore_index=True)

    # 定义成本要素下多种类函数检查并返回结果
    def check_duplicate(row, group):
        group_subset = group[
            (group['年'] == row['年']) &
            (group['月份'] == row['月份']) &
            (group['成本对象'] == row['成本对象']) &
            (group['区域'] == row['区域']) &
            (group['辅料类别'] == row['辅料类别'])
        ]
        if len(group_subset) > 1:
            return "辅料类别下多种类"
        return None

    # 应用函数到每一行
    real_table['备注2'] = real_table.apply(check_duplicate, group=real_table, axis=1)

    # 保存调整后的数据
    real_table.to_excel(get_file_path(f"{m}月真实消耗基础数据_调整.xlsx"), index=False)

    # 制作表2：筛选条件：备注2不为“辅料类别下多种类”，年为2024，月在1到11范围内
    filtered_df = real_table[(real_table['备注2'] != '辅料类别下多种类') & (real_table['年'].between(2022, 2024)) & (real_table['月份'].between(1, 11))]

    # 计算同一区域、同一成本对象、同一存货编码下折算单耗的中位数
    mean_df = filtered_df.groupby(['区域', '成本对象', '存货编码'])['折算单耗'].median().reset_index()

    # 表2：年为2024，月为12
    real_table_2 = real_table[(real_table['年'] == 2025) & (real_table['月份'] == m)]

    # 将计算结果合并到表2
    real_table_2 = pd.merge(real_table_2, mean_df, on=['区域', '成本对象', '存货编码'], how='left', suffixes=('', '_折算单耗_平均'))

    # 保存结果表2到新的 Excel 文件
    with pd.ExcelWriter(get_file_path(f"{m}月真实消耗基础数据_调整.xlsx"), mode='a', engine='openpyxl') as writer:
        real_table_2.to_excel(writer, index=False, sheet_name='历史对比表')

    # 加载 Excel 文件
    wb = load_workbook(get_file_path(f"{m}月真实消耗基础数据_调整.xlsx"))
    # 获取“历史对比表”工作表
    ws = wb['历史对比表']

    # 删除第 24 列（X 列）及其右侧的所有列
    max_col = ws.max_column
    if max_col >= 24:
        ws.delete_cols(24, max_col - 24 + 1)

    # 保存修改后的 Excel 文件
    wb.save(get_file_path(f"{m}月真实消耗基础数据_调整.xlsx"))
    return m

# 三、处理烟用材料表格
def process_smoking_material_data():
    file_path = get_file_path("烟用材料消耗数据收集.xlsx")
    xls = pd.ExcelFile(file_path)

    # 处理每个子表，修改名称并清理数据
    processed_sheets = {}
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=1)
        df = df.iloc[:, :10]  # 删除 K 列（从第 10 列起的所有列）
        
        # 修改表格名称
        new_sheet_name = clean_sheet_name(sheet_name)
        new_sheet_name = further_modify_sheet_name(new_sheet_name)
        processed_sheets[new_sheet_name] = df

        # 修改'材料类别'列的值
        if '材料类别' in df.columns:
            df['材料类别'] = df['材料类别'].replace({'烟用滤棒': '滤棒', '盒包装材料': '盒包', '条包装材料': '条包'})

        # 处理拉线数据
        if '材料类别' in df.columns and '主计量单位实际消耗数' in df.columns:
            # 筛选出材料类别为“拉线 - 小盒”和“拉线 - 条盒”的数据
            拉线_subset = df[df['材料类别'].isin(['拉线-小盒', '拉线-条盒'])]
            if not 拉线_subset.empty:
                # 计算主计量单位实际消耗数的总和
                total_consumption = 拉线_subset['主计量单位实际消耗数'].sum()
                # 创建新行
                new_row = df.iloc[0].copy()  # 拷贝第一行
                new_row['材料类别'] = '拉线'  # 设置新行的'材料类别'为'拉线'
                new_row['主计量单位实际消耗数'] = total_consumption  # 设置新行的消耗数
                # 将新行添加到数据框中
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

        # 将处理后的数据添加到字典中
        processed_sheets[new_sheet_name] = df

    # 保存修改后的表格
    with pd.ExcelWriter(get_file_path('烟用材料消耗数据收集_调整.xlsx')) as writer:
        for sheet_name, df in processed_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)


# 四、填入数据（假设文件路径和需求类似）
def fill_in_data(m):
    # 读取文件并处理数据
    file_path = get_file_path(f"{m}月真实消耗基础数据_调整.xlsx")  # 获取文件路径，使用传入的 m
    excel_file = pd.ExcelFile(file_path)
    df = excel_file.parse('历史对比表', dtype={'存货编码': str})  # 读取 '历史对比表' 工作表
    
    # 第二步：对 '历史对比表' 进行预处理
    df['辅料类别'] = df['辅料类别'].str.strip()  # 去除辅料类别列中的空格
    df['区域'] = df['区域'].str.strip()  # 去除区域列的空格
    df['成本对象'] = df['成本对象'].str.strip()  # 去除成本对象列的空格
    df['存货编码'] = df['存货编码'].str.strip()  # 去除存货编码列的空格

    # 定义根据“存货编码”调整“辅料类别”的函数
    def adjust_material_category(row):
        if row['辅料类别'] == '透明包装膜':
            stock_code = str(row['存货编码'])
            if stock_code.startswith('031201'):
                return '盒包透明包装膜'
            elif stock_code.startswith('031202'):
                return '条包透明包装膜'
        return row['辅料类别']

    # 应用函数来调整“辅料类别”列
    df['辅料类别'] = df.apply(adjust_material_category, axis=1)

    # 找到“折算单耗”列的索引
    index_of_conversion_consumption = df.columns.get_loc('折算单耗')

    # 在“折算单耗”列的右边插入“合并单耗”列，初始值设为 None
    df.insert(index_of_conversion_consumption + 1, '合并单耗', None)

    # 找到“备注2”列的索引
    index_of_remark2 = df.columns.get_loc('备注2')

    # 在“备注2”列的右边插入“折算单耗_折算单耗_中位数”列，初始值设为 None
    df.insert(index_of_remark2 + 1, '折算单耗_中位数', None)

    # 第三步：读取其他文件并修改 '历史对比表'
    file_path_2 = get_file_path("烟用材料消耗数据收集_调整.xlsx")
    excel_file_2 = pd.ExcelFile(file_path_2)
    sheet_names_2 = excel_file_2.sheet_names

    # 遍历 '历史对比表' 的每一行
    for index_1, row_1 in df.iterrows():
        cost_object = row_1['成本对象']
        material_category = row_1['辅料类别']

        # 找到对应的sheet
        for sheet_name in sheet_names_2:
            if cost_object in sheet_name:
                df2 = excel_file_2.parse(sheet_name)

                # 在表2中寻找匹配的材料类别
                for index_2, row_2 in df2.iterrows():
                    if row_2['材料类别'] == material_category:
                        # 填写到 '折算单耗_折算单耗_中位数' 字段
                        df.at[index_1, '折算单耗_中位数'] = row_2['主计量单位实际消耗数']
                        break

    # 第四步：合并单耗计算
    # 按成本对象、区域和辅料类别分组，计算每组折算单耗的总和
    grouped = df.groupby(['成本对象', '区域', '辅料类别'])['折算单耗'].sum()

    # 遍历数据框的每一行，计算合并单耗
    for index, row in df.iterrows():
        cost_object = row['成本对象']
        region = row['区域']
        material_category = row['辅料类别']
        # 获取合并单耗值
        df.at[index, '合并单耗'] = grouped[(cost_object, region, material_category)]

    # 第五步：获取丝束和成型纸的折算单耗_中位数
    final_file_path = get_file_path("真实消耗基础数据.xlsx" )
    final_excel_file = pd.ExcelFile(final_file_path)
    full_table = final_excel_file.parse('完整表', dtype={'存货编码': str})

    # 清理数据中的空格
    full_table['区域'] = full_table['区域'].str.strip()
    full_table['成本对象'] = full_table['成本对象'].str.strip()
    full_table['存货编码'] = full_table['存货编码'].str.strip()

    # 遍历历史对比表的每一行
    for index, row in df.iterrows():
        if row['辅料类别'] in ['丝束', '成型纸']:
            region = row['区域']
            cost_object = row['成本对象']
            stock_code = row['存货编码']
            # 从完整表中筛选出相同区域、成本对象和存货编码的数据
            filtered_data = full_table[
                (full_table['区域'] == region) &
                (full_table['成本对象'] == cost_object) &
                (full_table['存货编码'] == stock_code)
            ]
            if not filtered_data.empty:
                # 计算折算单耗的中位数
                median_value = filtered_data['折算单耗'].median()
                df.at[index, '折算单耗_中位数'] = median_value

    # 第六步：获取卷烟纸的折算单耗_中位数
    for index, row in df.iterrows():
        if row['辅料类别'] == '卷烟纸':
            region = row['区域']
            cost_object = row['成本对象']
            stock_code = row['存货编码']
            # 从完整表中筛选出相同区域、成本对象和存货编码的数据
            filtered_data = full_table[
                (full_table['区域'] == region) &
                (full_table['成本对象'] == cost_object) &
                (full_table['存货编码'] == stock_code)
            ]
            # 过滤备注2为"辅料类别下多种类"的行
            filtered_data = filtered_data[filtered_data['备注2'] != '辅料类别下多种类']
            if not filtered_data.empty:
                # 计算折算单耗的中位数
                median_value = filtered_data['折算单耗'].median()
                df.at[index, '折算单耗_中位数'] = median_value

    # 第七步之前新增功能：计算并插入“差值百分比(与合并单耗)”列df['合并单耗']
    df['差值百分比[(合并单耗/折算单耗_中位数)-1]'] = ((df['合并单耗'] / df['折算单耗_中位数']) - 1) * 100  # 计算差值百分比
    df['差值百分比[(合并单耗/折算单耗_中位数)-1]'] = df['差值百分比[(合并单耗/折算单耗_中位数)-1]'].apply(lambda x: f"{x:.2f}%")  # 将百分比格式化为百分号形式
    # 第七步：保存结果
    output_path = get_file_path(f"{m}月真实消耗基础数据_最终.xlsx")  # 使用 m 作为文件名的一部分
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 获取原文件所有表名
        sheet_names = excel_file.sheet_names
        for sheet in sheet_names:
            if sheet == '历史对比表':
                # 写入修改后的历史对比表
                df.to_excel(writer, sheet_name=sheet, index=False)
            else:
                # 写入其他未修改的表
                df_original = excel_file.parse(sheet)
                df_original.to_excel(writer, sheet_name=sheet, index=False)


# 清理表格名字的函数
def clean_sheet_name(sheet_name):
    if "七匹狼" in sheet_name or "万宝路" in sheet_name or "软富健" in sheet_name or "古田" in sheet_name:
        sheet_name = sheet_name.replace("龙岩", "").replace("，", "").replace("（", "(").replace("）", ")")
    return sheet_name

# 进一步修改子表名称的函数
def further_modify_sheet_name(sheet_name):
    replacements = {
        "七匹狼(尚品软包硬化)": "七匹狼(尚品)",
        "七匹狼(1575二维)": "七匹狼(1575)",
        "七匹狼(成功细支)": "七匹狼(古田成功细支)",
        "七匹狼(古田金细支二维)": "七匹狼(古田金细支)",
        "七匹狼(古田金中支二维)": "七匹狼(古田金中支)",
        "七匹狼(鼓浪扬帆二维)": "七匹狼(鼓浪扬帆)",
        "七匹狼(观海中支二维)": "七匹狼(观海中支)",
        "七匹狼(银中支二维)": "七匹狼(银中支)",
        "七匹狼(英伦奶香二维)": "七匹狼(英伦奶香)"
    }
    return replacements.get(sheet_name, sheet_name)

# 执行任务（根据需求选择执行哪个功能）
if __name__ == "__main__":
    generate_consumption_check_table()
    m = adjust_real_consumption_data()
    process_smoking_material_data()
    fill_in_data(m)
    show_personal_info()
