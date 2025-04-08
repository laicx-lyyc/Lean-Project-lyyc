import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# 获取用户输入的区号m
m = input("请输入区号m：")

# 获取当前代码文件的路径
current_folder = os.path.dirname(os.path.realpath(__file__))

# 构造文件的完整路径
file_path = os.path.join(current_folder, f"{m}区辅料数据.xlsx")

# 读取文件
df = pd.read_excel(file_path, header=None)

# 删除第一行 (序号为0)
df = df.drop(0)

# 设置第二行（序号为1）为表头
df.columns = df.iloc[0]
df = df.drop(1)

# 根据输入的m来决定操作的字段
if m == '1':
    business_types = ['机台领料', '机台退料']
    business_columns = ['机台领料', '机台退料']
    check_columns = ['机台领料', '机台退料']
elif m == '2':
    business_types = ['车间暂存区发料到机台', '机台人工退车间暂存区']
    business_columns = ['车间暂存区发料到机台', '机台人工退车间暂存区']
    check_columns = ['车间暂存区发料到机台', '机台人工退车间暂存区']
else:
    print("输入无效，请输入1或2。")
    exit()

# 筛选出相关的业务类型数据
df_filtered = df[df['业务类型'].isin(business_types)]

# 检查是否有对应的业务类型
has_business_type = {column: column in df_filtered['业务类型'].values for column in check_columns}

# 对数据按“班级”、“辅料名称”、“交班时间”、“出入库标志”、"机台号"和"单据号"分组，计算相关字段的总和
df_grouped = df_filtered.groupby(['班级', '辅料名称', '交班时间', '出入库标志', '机台号', '单据号', '业务类型']).agg({'总数量': 'sum'}).reset_index()

# 创建“对账”子表，添加“机台号”和“单据号”列
df_reconciliation = pd.pivot_table(df_grouped, index=['班级', '辅料名称', '交班时间', '出入库标志', '机台号', '单据号'], 
                                  columns='业务类型', values='总数量', aggfunc='sum').reset_index()

# 检查是否有相关的列，如果没有，则添加并设置为None
for column in business_columns:
    if column not in df_reconciliation.columns:
        df_reconciliation[column] = None

# 计算每个“班级”和“辅料名称”对应的总和
df_totals = df_reconciliation.groupby(['班级', '辅料名称']).agg({col: 'sum' for col in business_columns}).reset_index()

# 将总和数据添加到“对账”子表的每个“班级”和“辅料名称”数据下面
df_final = pd.DataFrame()
for name, group in df_reconciliation.groupby(['班级', '辅料名称']):
    df_final = pd.concat([df_final, group])
    
    # 复制df_totals行，避免警告
    totals_row = df_totals[(df_totals['班级'] == name[0]) & (df_totals['辅料名称'] == name[1])].copy()  # .copy() 显式创建副本
    totals_row.loc[:, '机台号'] = '总和'  # 使用 .loc 来修改数据
    totals_row.loc[:, '单据号'] = ''  # 使用 .loc 来修改数据

    # 填充相关的业务列，如果没有数据，则为空或0
    for column in business_columns:
        if column in totals_row.columns:
            totals_row.loc[:, column] = totals_row[column] if has_business_type[column] else None
    
    df_final = pd.concat([df_final, totals_row])

# 保存数据到新的Excel文件
output_path = os.path.join(current_folder, f"{m}区处理后的辅料数据.xlsx")
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='原始数据', index=False)
    df_final.to_excel(writer, sheet_name='对账', index=False)

# 加载刚刚保存的Excel文件进行样式设置
wb = load_workbook(output_path)
ws = wb['对账']

# 设置字体加粗并为字体设置红色
font_red_bold = Font(bold=True, color="FF0000")

# 对总和行（每个班级和辅料名称的最后一行）应用红色加粗字体
for row_num, row in enumerate(df_final.values, 2):  # Excel行数从1开始，header在第一行
    if row[4] == '总和':  # 检查“机台号”列是否为“总和”
        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = font_red_bold

# 保存最终的Excel文件
wb.save(output_path)

print(f"操作完成，文件已保存：{output_path}")
